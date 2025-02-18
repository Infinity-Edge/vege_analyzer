import pandas as pd
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os
import openpyxl

# Load the Excel file
def load_excel(file_path):
    return pd.ExcelFile(file_path)

# Process the file and calculate price changes
def process_price_changes(file_path, previous_price_file, output_path):
    # Load the files
    xls = load_excel(file_path)
    previous_prices = pd.read_csv(previous_price_file)

    # Load the necessary sheets into separate dataframes
    current_prices = pd.read_excel(xls, sheet_name='단가입력')
    standard_specs = pd.read_excel(xls, sheet_name='기준규격')

    # Prepare an empty list for unit prices
    unit_prices = []

    # Populate unit_prices by checking current_prices and standard_specs
    for _, row in current_prices.iterrows():
        품목명 = row['품목명']
        단위 = row['기준단위']
        단가 = row['단가']

        # Append the original entry
        unit_prices.append({'품목명': 품목명, '단위': 단위, '단가': math.ceil(단가 / 100) * 100})

        # Check for other units in standard_specs
        if 품목명 in standard_specs['품목명'].values:
            spec_row = standard_specs[standard_specs['품목명'] == 품목명]

            if '키로_단위' in spec_row and not pd.isna(spec_row['키로_단위'].values[0]):
                unit_prices.append(
                    {'품목명': 품목명, '단위': '키로', '단가': math.ceil((단가 / spec_row['키로_단위'].values[0]) / 100) * 100}
                )

            if '낱개_단위' in spec_row and not pd.isna(spec_row['낱개_단위'].values[0]):
                unit_prices.append(
                    {'품목명': 품목명, '단위': '개', '단가': math.ceil((단가 / spec_row['낱개_단위'].values[0]) / 100) * 100}
                )

    # Convert unit_prices to a DataFrame
    unit_prices_df = pd.DataFrame(unit_prices)

    # Add previous price information for comparison
    price_comparison = pd.merge(unit_prices_df, previous_prices, on=['품목명', '단위'], how='left', suffixes=('', '_전날'))

    # Calculate price changes
    def calculate_change(row):
        if pd.isna(row['단가_전날']):
            return '전날단가X'
        change = row['단가'] - row['단가_전날']
        change_pct = (change / row['단가_전날']) * 100
        return f"{int(change)} ({int(change_pct)}%)"

    price_comparison['등락'] = price_comparison.apply(calculate_change, axis=1)

    # Create the output dataframe with specific columns
    output_data = price_comparison[['품목명', '단위', '단가', '단가_전날', '등락']]

    # Save the output to a new Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        output_data.to_excel(writer, sheet_name='일일단가', index=False)

    # Open the Excel file to apply formatting
    workbook = load_workbook(output_path)
    sheet = workbook['일일단가']

    # Adjust column widths and alignments
    sheet.column_dimensions['A'].width = 130 / 7.5  # Convert pixels to Excel width units
    sheet.column_dimensions['C'].width = 80 / 7.5
    sheet.column_dimensions['D'].width = 80 / 7.5
    sheet.column_dimensions['E'].width = 80 / 7.5

    # Center align the first row
    for cell in sheet[1]:
        cell.alignment = Alignment(horizontal='center')

    # Adjust alignment for other rows
    for col in ['D', 'E']:
        for cell in sheet[col]:
            if cell.row > 1:
                cell.alignment = Alignment(horizontal='right')

    # Adjust column width for '등락'
    max_width = 0
    for cell in sheet['E']:
        if cell.row > 1 and cell.value:
            max_width = max(max_width, len(str(cell.value)))
    sheet.column_dimensions['E'].width = max_width + 2  # Add padding for readability

    # Apply conditional formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, str) and '(' in cell.value:
                change_pct = int(cell.value.split('(')[1].replace('%', '').replace(')', ''))
                if change_pct >= 20:
                    cell.font = Font(color="FF0000", bold=True)
                elif change_pct <= -20:
                    cell.font = Font(color="0000FF", bold=True)

    workbook.save(output_path)

    return output_path

# Backup and update previous_price.csv
def update_previous_prices(price_changes_file, previous_price_file):
    # Backup existing previous_price.csv
    if os.path.exists(previous_price_file):
        backup_file = previous_price_file.replace(".csv", "_backup.csv")
        os.rename(previous_price_file, backup_file)

    # Load price_changes and extract necessary columns
    price_changes = pd.read_excel(price_changes_file, sheet_name='일일단가')
    updated_previous_prices = price_changes[['품목명', '단위', '단가']]

    # Save the updated previous prices to previous_price.csv
    updated_previous_prices.to_csv(previous_price_file, index=False)
    print(f"Updated {previous_price_file} and created backup as {backup_file}")



# Function to classify and split item details
def classify_item_details(string):
    # Initialize result dictionary
    result = {'품목': '', '규격': '', '수량': '', '단위': '', '처리': ''}

    # Check if the string contains a space
    if " " not in string:
        result['품목'] = string
        result['처리'] = '기타'
        return result

    # Split the string into A_part and B_part (from the last space)
    last_space_index = string.rfind(" ")
    a_part = string[:last_space_index]
    b_part = string[last_space_index + 1:]

    # Process A_part
    if "(" in a_part and ")" in a_part:
        result['규격'] = a_part[a_part.index("("):a_part.index(")") + 1]
        result['품목'] = a_part[:a_part.index("(")].strip()
    else:
        result['품목'] = a_part

    # Process B_part
    if "(" in b_part and ")" in b_part:
        result['처리'] = b_part[b_part.index("("):b_part.index(")") + 1]

    # Extract quantity and unit from B_part
    import re
    quantity_match = re.search(r'[0-9+.]+', b_part)
    if quantity_match:
        result['수량'] = quantity_match.group()
        unit_start = quantity_match.end()
        unit_end = b_part.index("(") if "(" in b_part else len(b_part)
        result['단위'] = b_part[unit_start:unit_end].strip()

    return result

# Function to process the workbook
def process_workbook(file_paths, price_changes_path):
    # Load price changes
    price_changes = pd.read_excel(price_changes_path)
    price_map = price_changes.set_index(['품목명', '단위'])['단가'].to_dict()

    # Process each file in the file_paths list
    for file_path in file_paths:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)

        # Get all sheet names
        sheet_names = workbook.sheetnames

        # Sheets to exclude
        exclude_sheets = ['한울', '공산', '단가기입', '전날단가']

        # Process each sheet except the excluded ones
        for sheet_name in sheet_names:
            if sheet_name in exclude_sheets:
                continue

            sheet = workbook[sheet_name]

            # Check if the sheet is a worksheet
            if not isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                continue

            # Iterate through column A
            for row in sheet.iter_rows(min_col=1, max_col=1, min_row=1, max_row=sheet.max_row):
                cell = row[0]
                if cell.value and isinstance(cell.value, str):
                    cell_value = cell.value.strip()
                    details = classify_item_details(cell_value)

                    # Check for matching price in price map
                    key = (details['품목'], details['단위'])
                    if key in price_map:
                        details['단가'] = price_map[key]
                        print(f"File {file_path} - Sheet {sheet_name} - Row {cell.row}: {details}")
                    else:
                        print(f"File {file_path} - Sheet {sheet_name} - Row {cell.row}: {details}")


# Example usage
file_paths = ['/mnt/data/1.10장부_2컴.xlsm']  # 파일 리스트
price_changes_path = './price_changes.xlsx'
process_workbook(file_paths, price_changes_path)

print(f"Workbook processed and saved: {file_paths}")

'''
# Example usage
input_file = './product_spec_example.xlsx'  # Replace with your actual file path
previous_price_file = './previous_price.csv'  # Replace with your actual file path
output_file = './price_changes.xlsx'
process_price_changes(input_file, previous_price_file, output_file)

# Update previous prices
update_previous_prices(output_file, previous_price_file)

print(f"Price changes calculated, saved to {output_file}, and previous prices updated.")

print(f"Price changes calculated and saved to {output_file}")
'''

